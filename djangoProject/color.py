from openpyxl import Workbook
from django.http import StreamingHttpResponse
from openpyxl.styles import *
import pandas as pd
import os
import datetime
import uuid
class Mos_Process(object):
    """
    # MOS报表
    """

    def mos_tall(self, sheet, val, zm, i, j) -> None:
        # 判断前三列是否为三空
        _b1 = sheet[f"{mos_dic['excess']}" + str(j)].value
        _b2 = sheet[f"{mos_dic['age_180_plus']}" + str(j)].value
        _b3 = sheet[f"{mos_dic['age_91_to_180']}" + str(j)].value
        if not (_b1 == 0 and _b2 == 0 and _b3 == 0) and val == "MOS过高":
            # 红色高亮
            fill_1 = PatternFill("solid", fgColor="ff0000")
            sheet[f"{zm[i]}" + str(j)].fill = fill_1

    # 库存大于10
    def stock_gt_ten(self, sheet, item, i, j, zm, title) -> None:
        if sheet[f"{mos_dic['库存合计']}" + str(j)].value > 10:  # 判断库存是否大于10
            if item[title[i]] > 5:  # 判断当前列值是否大于5 红色
                fill_1 = PatternFill("solid", fgColor="ff0000")
                sheet[f"{zm[i]}" + str(j)].fill = fill_1
            elif item[title[i]] > 3:  # 判断当前列值是否大于3 浅红色
                fill_1 = PatternFill("solid", fgColor="C46565")  # C46565
                sheet[f"{zm[i]}" + str(j)].fill = fill_1

    # 价格标记
    def price_flag(self, sheet, item, i, j, zm, title) -> None:
        from decimal import Decimal
        if pd.isna(sheet[f"{mos_dic['MinPrice']}" + str(j)].value) or pd.isna(item['sale_price']):
            return  # 跳过空值
        else:
            try:
                minPrice = Decimal(str(sheet[f"{mos_dic['MinPrice']}" + str(j)].value))
                sale_price = Decimal(str(item['sale_price']))

                if sale_price < (minPrice * Decimal("0.98")):  # “sale_price”列小于“MinPrice”列的98%，标记红色
                    fill_1 = PatternFill("solid", fgColor="ff0000")
                    sheet[f"{mos_dic['sale_price']}" + str(j)].fill = fill_1

                if sale_price > minPrice + (minPrice * Decimal("0.2")):  # 如果大于“MinPrice”20%，标记绿色
                    fill_1 = PatternFill("solid", fgColor="009900")
                    sheet[f"{mos_dic['sale_price']}" + str(j)].fill = fill_1
            except Exception as e:
                print(e)

    # 调整月销量标记
    def sales_flag(self, sheet, item, i, j, zm, title) -> None:
        if sheet[f"{mos_dic['调整月销量']}" + str(j)].value > 10:
            if item[title[i]] < 1:  # 数字小于1，标记绿色
                fill_1 = PatternFill("solid", fgColor="009900")
                sheet[f"{zm[i]}" + str(j)].fill = fill_1
            elif item[title[i]] < 2:  # 只要数字小于2，标记浅绿色
                fill_1 = PatternFill("solid", fgColor="88CE88")
                sheet[f"{zm[i]}" + str(j)].fill = fill_1

Inv_arr = ["MSKU","公司SKU","数据","父体","X备注","备注","分割1","颜色大类",
                                 "细分颜色","组合方式","材质","可扩展",
                                 "20寸的尺寸","ASIN","分割2","IN-STOCK库存","FBA在库预估",
                                 "亚马逊总库存","3","海外仓库存","预创货件","4","7天流量","7天转化率",
                                 "7天订单数量","7天实际日均","常规补货计划日均","5","14天订单数量",
                                 "14天日均数量","7/14增长","14天流量","14天转化率","6","佣金","FBA费",
                                 "成本","高","长","宽","计费重","$头程","价格","定价毛利率","DEAL",
            "Coupon","Prime","OFF","售价","毛利率","毛利额$","7","定价毛利率",
           "INSTOCK库存售罄天数","FBA预估库存售罄天数",
           "FBA总库存（含在途）售罄天数","INSTOCK差额＜60",
           "实际售罄天数","8","9","安全天数","应补货天数","10",
           "15天特批","25天","35天","45天","60天","复核天数",
           "合计天数","差值","11","工厂待出货数量","预下单量","12","x备注","F"]

mos_dic = {}

def mos_wirte_excel(data, title, col=None, val=0, dirs="excel", fgColor="ff0000", is_exec=False) -> dict:
    global mos_dic
    # 清内存
    import gc
    wb = 0
    del wb
    gc.collect()
    """
    :param data     # 写入excel的数据
    :param title    # excel第一行标题
    :param col  # 比较第几列  高亮行数
    :param val  # 比较数字  必须整数
    :param dirs  # 保存文件子目录名
    :param fgColor  # 保存文件子目录名
    :return:
    """
    data = pd.DataFrame(data)[Inv_arr].to_dict("records")
    title = Inv_arr
    mp = Mos_Process()
    print(title)
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"

    # 表格冻结
    sheet.freeze_panes = "AA2"  # 冻结单元格所设置的参数为一个单元格，这个单元格上侧和左侧的所有行 / 列会被冻结
    wb.save(r'C:\Users\wb\Desktop\LXGG.xlsx')
    # sheet.freeze_panes = "A2"  # 第一行
    # 大写字母
    zm = [chr(i) for i in range(ord("A"), ord("Z") + 1)]
    # 遍历标题
    for t in title:
        if title.index(t) == 26:
            zm.extend(["A%s" % i for i in zm])
        elif title.index(t) == 52:
            zm.extend(["B%s" % i for i in zm])
        sheet[f"{zm[title.index(t)]}1"].value = t
        mos_dic.update({t: zm[title.index(t)]})  # {"标题": 列序号}
    j = 2

    for item in data:
        for i in range(len(title)):
            sheet[f"{zm[i]}" + str(j)].value = item[title[i]]  # 当前列值
            # if flag:
            #     sheet[f"{zm[i]}" + str(j)].fill = fill_1
            # if zm[i] == mos_dic["MOS高判断"]:  # MOS过高高亮
            #     mp.mos_tall(sheet, item[title[i]], zm, i, j)
            if zm[i] in [mos_dic["30天MOS(仅库存)"], mos_dic["14天MOS(仅库存)"], mos_dic["7天MOS(仅库存)"],
                         mos_dic["调整月销量MOS(仅库存)"]]:  # 库存合计
                """
                    在“库存合计”列大于10情况下，只要数字大于3，标记浅红色，数字大于5，标记红色
                """
                mp.stock_gt_ten(sheet, item, i, j, zm, title)
            if zm[i] in [mos_dic["7天MOS(含在途)"], mos_dic["调整月销量MOS(含在途)"], mos_dic["加上本地库存调整月销量MOS"]]:
                """
                这三列，在“调整月销量”列大于10情况下，只要数字小于2，标记浅绿色，数字小于1，标记绿色
                """
                mp.sales_flag(sheet, item, i, j, zm, title)

            if zm[i] == mos_dic["新品上架"]:
                """
                “sale_price”列小于“MinPrice”列的98%，标记红色；如果大于“MinPrice”20%，标记绿色
                """
                mp.price_flag(sheet, item, i, j, zm, title)
        j += 1


    # S1设置成黄色
    fill_1 = PatternFill("solid", fgColor="FFFF00")
    sheet[f"{mos_dic['7天MOS(含在途)']}1"].fill = fill_1

    # 换行设置
    ali = Alignment()
    ali.wrapText = 1
    for i in range(1, 36):
        sheet[zm[i] + "1"].alignment = ali


    # 清内存
    import gc
    del wb  # wb为打开的工作表
    gc.collect()
    return {"code": 1, "msg": "ok", "data": {"save_path": response_path}}